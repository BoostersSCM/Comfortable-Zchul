# app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
import urllib.parse

# =========================================================
# 0. 페이지 설정
# =========================================================
st.set_page_config(page_title="Boosters 납품대금 집계 시스템", layout="wide")

# =========================================================
# 1. OAuth Secrets
# =========================================================
try:
    CLIENT_ID = st.secrets["google_auth"]["client_id"]
    CLIENT_SECRET = st.secrets["google_auth"]["client_secret"]
    REDIRECT_URI = st.secrets["google_auth"]["redirect_uri"]
except Exception:
    st.error("⚠️ Secrets 설정이 되어있지 않습니다. Streamlit Cloud의 Settings > Secrets를 확인해주세요.")
    st.stop()

# =========================================================
# 2. Google OAuth 함수
# =========================================================
def get_login_url():
    base_url = "https://accounts.google.com/o/oauth2/v2/auth"
    params = {
        "response_type": "code",
        "client_id": CLIENT_ID,
        "redirect_uri": REDIRECT_URI,
        "scope": "openid email profile",
        "access_type": "offline",
        "prompt": "consent",
        "hd": "boosters.kr",
    }
    return f"{base_url}?{urllib.parse.urlencode(params)}"

def get_token_from_code(code: str) -> dict:
    token_url = "https://oauth2.googleapis.com/token"
    data = {
        "code": code,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "redirect_uri": REDIRECT_URI,
        "grant_type": "authorization_code",
    }
    return requests.post(token_url, data=data, timeout=20).json()

def get_user_info(access_token: str) -> dict:
    user_info_url = "https://openidconnect.googleapis.com/v1/userinfo"
    headers = {"Authorization": f"Bearer {access_token}"}
    return requests.get(user_info_url, headers=headers, timeout=20).json()

# =========================================================
# 3. Query params 호환 레이어 (중요)
# =========================================================
def get_all_query_params() -> dict:
    """
    Streamlit 버전에 따라 query params API가 달라서 둘 다 지원.
    반환 형태는 dict[str, list[str]] 로 통일.
    """
    # 최신: st.query_params
    if hasattr(st, "query_params"):
        qp_obj = st.query_params
        # QueryParams는 dict처럼 동작하지만, to_dict가 없는 버전도 있어서 안전하게 변환
        out = {}
        try:
            for k in qp_obj.keys():
                v = qp_obj.get(k)
                if isinstance(v, list):
                    out[k] = v
                elif v is None:
                    out[k] = []
                else:
                    out[k] = [str(v)]
            return out
        except Exception:
            pass

    # 구버전: st.experimental_get_query_params
    if hasattr(st, "experimental_get_query_params"):
        qp = st.experimental_get_query_params()
        # 이미 dict[str, list[str]]
        return {k: [str(x) for x in v] for k, v in qp.items()}

    return {}

def qp_first(key: str):
    qp = get_all_query_params()
    v = qp.get(key, [])
    return v[0] if v else None

def clear_query_params():
    # 최신
    if hasattr(st, "query_params"):
        try:
            st.query_params.clear()
            return
        except Exception:
            pass
    # 구버전
    if hasattr(st, "experimental_set_query_params"):
        st.experimental_set_query_params()

# =========================================================
# 4. 로그인 유지(쿠키)
# =========================================================
COOKIE_EMAIL = "boosters_login"
COOKIE_EXPIRY = "boosters_login_expiry"
COOKIE_DAYS = 7

def cookies_supported() -> bool:
    return hasattr(st, "cookies")

def set_login_cookie(email: str, days: int = COOKIE_DAYS):
    if not cookies_supported():
        return
    expire_at = datetime.utcnow() + timedelta(days=days)
    st.cookies[COOKIE_EMAIL] = email
    st.cookies[COOKIE_EXPIRY] = expire_at.isoformat()

def clear_login_cookie():
    if not cookies_supported():
        return
    st.cookies.pop(COOKIE_EMAIL, None)
    st.cookies.pop(COOKIE_EXPIRY, None)

def restore_login_from_cookie() -> bool:
    if not cookies_supported():
        return False
    email = st.cookies.get(COOKIE_EMAIL)
    expiry = st.cookies.get(COOKIE_EXPIRY)
    if not email or not expiry:
        return False
    try:
        if datetime.utcnow() < datetime.fromisoformat(expiry) and email.endswith("@boosters.kr"):
            st.session_state.logged_in = True
            st.session_state.user_email = email
            return True
    except Exception:
        return False
    return False

# =========================================================
# 5. 파일 읽기(헤더행 선택)
# =========================================================
def read_file_with_header(uploaded_file, header_row_excel_1based: int, header_row_csv_1based: int = 1):
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        header_idx = max(header_row_csv_1based - 1, 0)
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, header=header_idx)
        except Exception:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, header=header_idx, encoding="cp949")
        return df

    header_idx = max(header_row_excel_1based - 1, 0)
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=header_idx, engine="openpyxl")
    return df

# =========================================================
# 6. ERP 데이터 처리
# =========================================================
def load_and_aggregate_data(uploaded_file, header_row_excel_1based: int):
    try:
        df = read_file_with_header(uploaded_file, header_row_excel_1based=header_row_excel_1based)
    except Exception as e:
        return None, f"파일 읽기 실패: {e}"

    df.columns = [str(col).strip() for col in df.columns]
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]

    column_mapping = {
        "거래처": "업체",
        "발주번호": "발주번호",
        "품번": "품번",
        "품명": "품명",
        "단가": "납품단가",
        "납품수량": "납품수량",
        "금액": "납품금액(세전)",
        "부가세": "부가세",
        "금액계": "납품금액(세후)",
    }

    valid_cols = [col for col in column_mapping.keys() if col in df.columns]
    if not valid_cols:
        return None, f"필수 컬럼 없음. 감지된 제목: {list(df.columns)}"

    df_extracted = df[valid_cols].copy()
    df_extracted.rename(columns=column_mapping, inplace=True)

    numeric_cols = ["납품수량", "납품금액(세전)", "부가세", "납품금액(세후)"]
    for col in numeric_cols:
        df_extracted[col] = pd.to_numeric(
            df_extracted[col].astype(str).str.replace(",", ""),
            errors="coerce",
        ).fillna(0)

    group_keys = ["업체", "발주번호", "품번", "품명"]
    df_grouped = df_extracted.groupby(group_keys, as_index=False)[
        ["납품수량", "납품금액(세전)", "부가세", "납품금액(세후)"]
    ].sum()

    df_grouped["납품단가"] = df_grouped.apply(
        lambda x: x["납품금액(세전)"] / x["납품수량"] if x["납품수량"] != 0 else 0,
        axis=1,
    )

    df_grouped = df_grouped.sort_values(by=["업체", "발주번호", "품번"])

    desired_order = [
        "업체", "발주번호", "품번", "품명",
        "납품단가", "납품수량", "납품금액(세전)", "부가세", "납품금액(세후)"
    ]
    df_final = df_grouped[desired_order].copy()
    df_final["선금 지급일"] = ""
    df_final["선금 금액"] = 0
    df_final["잔여금액"] = 0

    return df_final, None

def create_excel_with_formula(df: pd.DataFrame) -> BytesIO:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    header_map = {str(cell.value).strip(): cell.col_idx for cell in ws[1]}
    needed = {"납품금액(세후)", "선금 금액", "잔여금액"}
    if needed.issubset(set(header_map.keys())):
        col_total = get_column_letter(header_map["납품금액(세후)"])
        col_prepay = get_column_letter(header_map["선금 금액"])
        col_balance = get_column_letter(header_map["잔여금액"])

        for r in range(2, ws.max_row + 1):
            ws[f"{col_balance}{r}"] = f"={col_total}{r}-{col_prepay}{r}"
            cols_to_format = ["납품단가", "납품금액(세전)", "부가세", "납품금액(세후)", "선금 금액", "잔여금액"]
            for col_name in cols_to_format:
                if col_name in header_map:
                    ws[f"{get_column_letter(header_map[col_name])}{r}"].number_format = "#,##0"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# =========================================================
# 7. 화면 표시용 DF (Styler 미사용)
# =========================================================
def make_display_df(df: pd.DataFrame) -> pd.DataFrame:
    df_disp = df.copy()
    num_cols = ["납품단가", "납품수량", "납품금액(세전)", "부가세", "납품금액(세후)", "선금 금액", "잔여금액"]
    for c in num_cols:
        if c in df_disp.columns:
            s = pd.to_numeric(df_disp[c], errors="coerce").fillna(0)
            df_disp[c] = s.map(lambda x: f"{int(round(x)):,}")
    return df_disp

# =========================================================
# 8. 메인 앱
# =========================================================
def main_app():
    with st.sidebar:
        st.success(f"접속자: {st.session_state.user_email}")
        if st.button("로그아웃"):
            st.session_state.clear()
            clear_login_cookie()
            st.rerun()
        st.caption(f"로그인 유지: {COOKIE_DAYS}일 (쿠키 기반)")

    st.title("📊 납품대금 집계 프로그램")

    uploaded_file = st.file_uploader("파일 업로드 (xlsx, xls, csv)", type=["xlsx", "xls", "csv"])

    header_row_excel = st.number_input("엑셀 헤더 행(1부터)", min_value=1, value=2, step=1)

    if "processed_data" not in st.session_state:
        st.session_state.processed_data = None

    if uploaded_file:
        with st.expander("🔎 헤더 미리보기(현재 설정 기준)", expanded=True):
            try:
                preview = read_file_with_header(uploaded_file, header_row_excel_1based=header_row_excel)
                st.write("감지된 컬럼:", list(preview.columns))
                st.dataframe(preview.head(5), use_container_width=True)
            except Exception as e:
                st.error(f"미리보기 실패: {e}")

        if st.button("🚀 변환 및 집계 실행", type="primary"):
            with st.spinner("데이터 분석 중..."):
                df_result, error_msg = load_and_aggregate_data(uploaded_file, header_row_excel_1based=header_row_excel)
                if df_result is not None:
                    st.session_state.processed_data = df_result
                    st.success("집계 완료!")
                else:
                    st.error(f"오류: {error_msg}")

    if st.session_state.processed_data is not None:
        st.divider()
        st.subheader("📋 결과 미리보기")
        st.dataframe(make_display_df(st.session_state.processed_data), use_container_width=True)

        excel_data = create_excel_with_formula(st.session_state.processed_data)
        st.download_button(
            label="📥 엑셀 파일 다운로드",
            data=excel_data,
            file_name="납품대금_집계표.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# =========================================================
# 9. 실행 흐름 제어 (로그인)
# =========================================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "user_email" not in st.session_state:
    st.session_state.user_email = ""

# 쿠키로 로그인 복원
if not st.session_state.logged_in:
    restore_login_from_cookie()

# ---- 디버그: 현재 쿼리 파라미터 확인 (로그인 안 될 때 매우 유용)
# 필요 없으면 주석 처리 가능
DEBUG_OAUTH = True

qp_all = get_all_query_params()
err = qp_first("error")
code = qp_first("code")

if DEBUG_OAUTH and (err or code):
    st.info("🔍 OAuth 디버그(현재 URL 쿼리 파라미터)")
    st.write(qp_all)

# OAuth error 처리
if err:
    st.error("Google OAuth 에러 발생")
    st.write(qp_all)
    st.stop()

# 로그인 처리
if not st.session_state.logged_in:
    if code:
        token_res = get_token_from_code(code)

        if DEBUG_OAUTH:
            st.info("🔍 Token response")
            st.write(token_res)

        if "access_token" not in token_res:
            st.error("로그인 실패: 토큰 발급 실패")
            st.stop()

        user_info = get_user_info(token_res["access_token"])
        if DEBUG_OAUTH:
            st.info("🔍 User info")
            st.write(user_info)

        email = user_info.get("email", "")

        if email.endswith("@boosters.kr"):
            st.session_state.logged_in = True
            st.session_state.user_email = email
            set_login_cookie(email, days=COOKIE_DAYS)

            clear_query_params()
            st.rerun()
        else:
            st.error(f"접속 권한이 없습니다. ({email}) @boosters.kr 계정만 가능합니다.")
            st.stop()

    else:
        st.title("🔒 Boosters Internal Tool")
        st.write("관계자 외 접근을 금지합니다.")
        login_url = get_login_url()
        st.markdown(
            f"""
            <a href="{login_url}" target="_self">
                <button style="
                    background-color: #4285F4; color: white; padding: 12px 24px;
                    border: none; border-radius: 6px; cursor: pointer;
                    font-size: 16px; font-weight: bold;">
                    G Suite 계정으로 로그인 (Boosters)
                </button>
            </a>
            """,
            unsafe_allow_html=True,
        )
        if not cookies_supported():
            st.warning("현재 Streamlit 버전에서 st.cookies가 지원되지 않아 '로그인 유지'가 동작하지 않을 수 있습니다.")
else:
    main_app()
